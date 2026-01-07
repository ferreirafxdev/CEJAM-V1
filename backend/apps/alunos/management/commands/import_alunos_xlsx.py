import json
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.core.exceptions import ValidationError
from django.utils.dateparse import parse_date

from apps.alunos.models import Aluno
from apps.professores.models import Professor
from apps.turmas.models import Turma

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise CommandError("openpyxl nao instalado. Instale com: pip install openpyxl") from exc


REQUIRED_COLUMNS = [
    "nome_completo",
    "data_nascimento",
    "sexo",
    "endereco",
    "telefone",
    "data_matricula",
    "turma",
    "valor_mensalidade",
]

OPTIONAL_COLUMNS = [
    "cpf",
    "responsavel_legado_nome",
    "responsavel_legado_telefone",
    "responsavel_legado_email",
    "status",
    "numero_matricula",
    "observacoes",
    "historico_escolar",
]

ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_digits(value, length=None):
    raw = normalize_text(value)
    if not raw:
        return ""
    raw = re.sub(r"\D", "", raw)
    if not raw:
        return ""
    if length and len(raw) < length:
        raw = raw.zfill(length)
    return raw


def normalize_matricula(value):
    if value is None:
        return ""
    if isinstance(value, (int, Decimal)):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    raw = normalize_text(value)
    if raw.endswith(".0") and raw.replace(".0", "").isdigit():
        raw = raw[:-2]
    return raw


def normalize_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    if isinstance(value, str):
        parsed = parse_date(value)
        if parsed:
            return parsed
        for fmt in ("%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def normalize_sexo(value):
    raw = normalize_text(value).upper()
    if raw in {"M", "MASCULINO"}:
        return "M"
    if raw in {"F", "FEMININO"}:
        return "F"
    if raw in {"O", "OUTRO"}:
        return "O"
    return ""


def normalize_status(value):
    raw = normalize_text(value).upper()
    if raw in {"ATIVO", "ATIVA"}:
        return "ATIVO"
    if raw in {"INATIVO", "INATIVA"}:
        return "INATIVO"
    return "ATIVO"


def normalize_decimal(value):
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    raw = normalize_text(value)
    raw = raw.replace(".", "").replace(",", ".")
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def parse_turma_map(raw_map):
    if not raw_map:
        return {}
    try:
        parsed = json.loads(raw_map)
    except json.JSONDecodeError as exc:
        raise CommandError("turma-map deve ser JSON valido.") from exc
    if not isinstance(parsed, dict):
        raise CommandError("turma-map deve ser um objeto JSON.")
    return {str(key).strip(): value for key, value in parsed.items()}


def resolve_turma(raw_value, turma_map, create_config, valor_mensalidade, cache):
    name = normalize_text(raw_value)
    if not name:
        return None
    mapped = turma_map.get(name, name)
    cache_key = str(mapped).strip()
    if cache_key in cache:
        return cache[cache_key]
    if isinstance(mapped, int) or (isinstance(mapped, str) and mapped.isdigit()):
        turma = Turma.objects.filter(id=int(mapped)).first()
        if turma:
            cache[cache_key] = turma
        return turma
    turma = Turma.objects.filter(nome__iexact=str(mapped)).first()
    if turma:
        cache[cache_key] = turma
        return turma
    if not create_config:
        return None
    professor = create_config["professor"]
    if not professor:
        return None
    serie_ano = create_config["serie_ano"] or str(mapped)
    turma = Turma(
        nome=str(mapped),
        serie_ano=serie_ano,
        turno=create_config["turno"],
        professor_responsavel=professor,
        valor_mensalidade=valor_mensalidade or create_config["valor_mensalidade"],
        capacidade_maxima=create_config["capacidade"],
        status=create_config["status"],
    )
    turma.full_clean()
    if not create_config["dry_run"]:
        turma.save()
    cache[cache_key] = turma
    return turma


class Command(BaseCommand):
    help = "Importa alunos de uma planilha XLSX."

    def add_arguments(self, parser):
        parser.add_argument("--path", required=True, help="Caminho do arquivo XLSX.")
        parser.add_argument("--sheet", default=None, help="Nome da aba (opcional).")
        parser.add_argument(
            "--match",
            choices=["cpf", "numero_matricula", "cpf_or_matricula"],
            default="cpf_or_matricula",
            help="Campo usado para localizar registros existentes.",
        )
        parser.add_argument(
            "--update-existing",
            action="store_true",
            help="Atualiza registros existentes em vez de ignorar.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simula a importacao sem gravar no banco.",
        )
        parser.add_argument(
            "--turma-map",
            default=None,
            help="JSON com mapeamento de turma. Ex: {\"Infantil I\": \"MATERNAL\"}",
        )
        parser.add_argument(
            "--create-turmas",
            action="store_true",
            help="Cria turmas inexistentes durante a importacao.",
        )
        parser.add_argument(
            "--turma-professor-id",
            type=int,
            default=None,
            help="ID do professor responsavel para turmas novas.",
        )
        parser.add_argument(
            "--turma-turno",
            choices=["MANHA", "TARDE", "NOITE"],
            default="MANHA",
            help="Turno padrão das turmas criadas.",
        )
        parser.add_argument(
            "--turma-capacidade",
            type=int,
            default=30,
            help="Capacidade maxima das turmas criadas.",
        )
        parser.add_argument(
            "--turma-status",
            choices=["ATIVA", "ENCERRADA"],
            default="ATIVA",
            help="Status das turmas criadas.",
        )
        parser.add_argument(
            "--turma-serie-ano",
            default=None,
            help="Serie/ano para turmas criadas (opcional).",
        )

    def handle(self, *args, **options):
        path = Path(options["path"]).expanduser()
        if not path.exists():
            raise CommandError(f"Arquivo nao encontrado: {path}")

        turma_map = parse_turma_map(options.get("turma_map"))
        create_config = None
        if options["create_turmas"]:
            professor_id = options.get("turma_professor_id")
            if not professor_id:
                raise CommandError("--turma-professor-id e obrigatorio com --create-turmas.")
            try:
                professor = Professor.objects.get(id=professor_id)
            except Professor.DoesNotExist as exc:
                raise CommandError("Professor nao encontrado para o ID informado.") from exc
            create_config = {
                "professor": professor,
                "turno": options["turma_turno"],
                "capacidade": options["turma_capacidade"],
                "status": options["turma_status"],
                "serie_ano": options.get("turma_serie_ano"),
                "valor_mensalidade": None,
                "dry_run": options["dry_run"],
            }

        wb = load_workbook(path, read_only=True, data_only=True)
        if options.get("sheet"):
            if options["sheet"] not in wb.sheetnames:
                raise CommandError(f"Aba nao encontrada: {options['sheet']}")
            ws = wb[options["sheet"]]
        else:
            ws = wb.active

        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            raise CommandError("Planilha vazia.")

        header_map = {normalize_text(col): idx for idx, col in enumerate(header) if col}
        missing = [col for col in REQUIRED_COLUMNS if col not in header_map]
        if missing:
            raise CommandError(f"Colunas obrigatorias ausentes: {', '.join(missing)}")

        total = 0
        created = 0
        updated = 0
        skipped = 0
        errors = []
        missing_turmas = set()
        turma_cache = {}

        for row_index, row in enumerate(rows, start=2):
            total += 1
            row_data = {}
            for col in ALL_COLUMNS:
                idx = header_map.get(col)
                row_data[col] = row[idx] if idx is not None and idx < len(row) else None

            nome_completo = normalize_text(row_data["nome_completo"])
            cpf = normalize_digits(row_data.get("cpf"), length=11) or None
            data_nascimento = normalize_date(row_data["data_nascimento"])
            sexo = normalize_sexo(row_data["sexo"])
            endereco = normalize_text(row_data["endereco"])
            telefone = normalize_text(row_data["telefone"])
            nome_responsavel = normalize_text(row_data.get("responsavel_legado_nome"))
            telefone_responsavel = normalize_text(row_data.get("responsavel_legado_telefone"))
            email_responsavel = normalize_text(row_data.get("responsavel_legado_email"))
            status = normalize_status(row_data.get("status"))
            data_matricula = normalize_date(row_data["data_matricula"])
            numero_matricula = normalize_matricula(row_data.get("numero_matricula")) or None
            turma_raw = row_data.get("turma")
            valor_mensalidade = normalize_decimal(row_data["valor_mensalidade"])
            observacoes = normalize_text(row_data.get("observacoes"))
            historico_escolar = normalize_text(row_data.get("historico_escolar"))

            if not nome_completo or not data_nascimento or not sexo or not endereco or not telefone:
                errors.append((row_index, "Campos obrigatorios ausentes."))
                skipped += 1
                continue

            if not data_matricula or valor_mensalidade is None:
                errors.append((row_index, "Data matricula ou valor invalido."))
                skipped += 1
                continue

            turma = resolve_turma(turma_raw, turma_map, create_config, valor_mensalidade, turma_cache)
            if not turma:
                missing_turmas.add(normalize_text(turma_raw))
                skipped += 1
                continue

            payload = {
                "nome_completo": nome_completo,
                "cpf": cpf,
                "data_nascimento": data_nascimento,
                "sexo": sexo,
                "endereco": endereco,
                "telefone": telefone,
                "responsavel": None,
                "nome_responsavel": nome_responsavel,
                "telefone_responsavel": telefone_responsavel,
                "email_responsavel": email_responsavel,
                "status": status,
                "data_matricula": data_matricula,
                "numero_matricula": numero_matricula,
                "turma": turma,
                "valor_mensalidade": valor_mensalidade,
                "observacoes": observacoes,
                "historico_escolar": historico_escolar,
            }

            existing = None
            match = options["match"]
            if match in {"cpf", "cpf_or_matricula"} and cpf:
                existing = Aluno.objects.filter(cpf=cpf).first()
            if not existing and match in {"numero_matricula", "cpf_or_matricula"} and numero_matricula:
                existing = Aluno.objects.filter(numero_matricula=numero_matricula).first()

            try:
                if existing:
                    if not options["update_existing"]:
                        skipped += 1
                        continue
                    for key, value in payload.items():
                        setattr(existing, key, value)
                    existing.full_clean()
                    if not options["dry_run"]:
                        existing.save()
                    updated += 1
                else:
                    aluno = Aluno(**payload)
                    aluno.full_clean()
                    if not options["dry_run"]:
                        aluno.save()
                    created += 1
            except ValidationError as exc:
                errors.append((row_index, str(exc)))
                skipped += 1

        self.stdout.write(f"Total linhas: {total}")
        self.stdout.write(f"Criados: {created}")
        self.stdout.write(f"Atualizados: {updated}")
        self.stdout.write(f"Ignorados: {skipped}")

        if missing_turmas:
            self.stdout.write("Turmas nao encontradas:")
            for name in sorted(missing_turmas):
                self.stdout.write(f" - {name}")

        if errors:
            self.stdout.write("Erros (primeiros 10):")
            for row_index, message in errors[:10]:
                self.stdout.write(f" - Linha {row_index}: {message}")
